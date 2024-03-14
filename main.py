# Importando as bibliotecas
import tkinter as tk
import tkinter.messagebox
from tkinter import Radiobutton, StringVar, messagebox
from openpyxl import load_workbook
from openpyxl import styles
from datetime import date

# Função que abre um messabox informando os dados do software
def Sobre():
  messagebox.showinfo("Sobre", "Direitos Autorais © 2024 Guilherme Lourenço Lopes. Todos os direitos reservados. Proibida a reprodução, distribuição ou modificação sem permissão expressa do autor.")
# Função que abre um messabox informando a versão dos software
def Versao():
  messagebox.showinfo("Versão", "Versão 2.0")
# Função responsável por realizar busca dentro da planilha excel
def Localizar():
  # Criação e configuração da janela
  localiza = tk.Toplevel()
  localiza.title("Localizar Rádio")
  localiza.geometry("400x50")
  localiza.configure(background='#a8b5ab')

  # Criação de variáveis
  Busca_OS_var = tk.StringVar()

  # Criando as label
  Busca_OS_label = tk.Label(localiza, text = "OS procurada:", bg='#8a948c', font = Font_3)
  
  # Posicionando os labels
  Busca_OS_label.grid(row=0, column=0, padx=10, pady=3)
  
  # Criando o campo de entrada
  Busca_OS_entry = tk.Entry(localiza, textvariable = Busca_OS_var, font = Font_3)

  # Posicionando o campo de entrada
  Busca_OS_entry.grid(row=0, column=1, sticky='E')

  # Caminho do arquivo Excel
  excel_file = 'MANUTENÇÕES INTERNAS WELLINGTON.xlsx'

  # Carrega o arquivo Excel
  wb = load_workbook(excel_file)
  planilha_radio = wb.active

  # Função para buscar e exibir informações
  def exibir_info():
    os_procurada = Busca_OS_var.get()
    linha_encontrada = None

    # Itera sobre as linhas da planilha para encontrar a OS
    for row in planilha_radio.iter_rows(min_row=2, max_row=planilha_radio.max_row, min_col=1, max_col=1):
      for cell in row:
        if str(cell.value) == os_procurada:
          linha_encontrada = cell.row
          break

    if linha_encontrada:
      # Obtém as informações da planilha
      PATRI = planilha_radio.cell(row=linha_encontrada, column=2).value
      OS_A = planilha_radio.cell(row=linha_encontrada, column=3).value
      GAR = planilha_radio.cell(row=linha_encontrada, column=4).value
      NS = planilha_radio.cell(row=linha_encontrada, column=5).value
      CLIEN = planilha_radio.cell(row=linha_encontrada, column=6).value
      EQUIPAMENTO = planilha_radio.cell(row=linha_encontrada, column=7).value
      STATUS = planilha_radio.cell(row=linha_encontrada, column=8).value
      IDENIZAR = planilha_radio.cell(row=linha_encontrada, column=9).value
      DATA = planilha_radio.cell(row=linha_encontrada, column=10).value
      RETORNO = planilha_radio.cell(row=linha_encontrada, column=11).value
      DEFEITO = planilha_radio.cell(row=linha_encontrada, column=12).value
      
      # Monta a mensagem para exibir na messagebox
      mensagem = f"OS Antiga: {OS_A}\nRetorno: {RETORNO}\nGARANTIA: {GAR}\nN/S: {NS}\nPatrimônio: {PATRI}\nProprietário: {CLIEN}\nEquipamento: {EQUIPAMENTO}\nStatus:{STATUS}\nIndenização: {IDENIZAR}\nData: {DATA}\nDefeito: {DEFEITO}"

      # Exibe a mensagem na messagebox
      messagebox.showinfo("Informações Encontradas", mensagem)
      localiza.destroy() 
    else:
      # Exibe uma mensagem se a OS não for encontrada
      messagebox.showinfo("Busca não encontrada", "O.S não encontrada")
      localiza.destroy() 

  # Botão para acionar a busca
  buscar_button = tk.Button(localiza, text='Buscar', command=exibir_info)
  buscar_button['background'] = '#8a948c'
  buscar_button['activebackground'] = 'gray40'
  buscar_button['fg'] = 'white'
  buscar_button['font'] = Font_1
  buscar_button.config(width=7, height=1)

  # Configurando a posição do botão submit
  buscar_button.place(relx=0.83, rely=0.3, anchor='center') 

  # Salvando a planilha
  wb.save(excel_file)
#Função responsável por alterar data de uma OS
def Alterar_Data():
  # Criação e configuração da janela
  altera = tk.Toplevel()
  altera.title("Alterar Data")
  altera.geometry("400x65")
  altera.configure(background='#a8b5ab')

  # Caminho do arquivo Excel
  excel_file = 'MANUTENÇÕES INTERNAS WELLINGTON.xlsx'

  # Carrega o arquivo Excel
  wb = load_workbook(excel_file)
  planilha_radio = wb.active
  
  #Declaração de variáveis
  OS_Alterar_var = tk.StringVar()
  Data_alterar_var = tk.StringVar()

  # Criando as label
  OS_Alterar_label = tk.Label(altera, text = "OS procurada:", bg='#8a948c', font = Font_3)
  Data_alterar_label = tk.Label(altera, text = "Alterar Data:", bg='#8a948c', font = Font_3)

  # Posicionando os labels
  OS_Alterar_label.grid(row=0, column=0, sticky='E', pady=3)
  Data_alterar_label.grid(row=2, column=0, sticky='E', pady=3)
  
  # Criando o campo de entrada
  OS_Alterar_entry = tk.Entry(altera, textvariable = OS_Alterar_var, font = Font_2)
  Data_alterar_entry = tk.Entry(altera, textvariable = Data_alterar_var, font = Font_2)

  # Posicionando o campo de entrada
  OS_Alterar_entry.grid(row=0, column=1, sticky='W')
  Data_alterar_entry.grid(row=2, column=1, sticky='W')

  def alterando_campos():
    os_procurada = OS_Alterar_var.get()
    linha_encontrada = None
  
    # Itera sobre as linhas da planilha para encontrar a OS
    for row in planilha_radio.iter_rows(min_row=2, max_row=planilha_radio.max_row, min_col=1, max_col=1):
      for cell in row:
        if str(cell.value) == os_procurada:
          linha_encontrada = cell.row
          break
    
    if linha_encontrada:
      nova_data = Data_alterar_var.get()

      c10 = planilha_radio.cell(row=linha_encontrada, column=10)

      # Apaga os valores das células
      c10.value = None

      # Atribuindo os novos valores às células
      c10.value = nova_data
    
      # Salvando a planilha
      wb.save(excel_file)

    else:
      # Mostrando uma mensagem de erro se a OS não for encontrada
      messagebox.showinfo("Busca não encontrada", "O.S não encontrada")
      altera.destroy()
      
    # Zerando as strings
    OS_Alterar_var.set("")
    Data_alterar_var.set("")
    
  # Configurando um botão para que os novos dados sejam inseridos na planilha do excel
  altera_button = tk.Button(altera ,text = 'Alterar', command = alterando_campos)
  altera_button['background'] = '#667369'
  altera_button['activebackground'] = 'gray40'
  altera_button['fg'] = 'white'
  altera_button['font'] = Font_1
  altera_button.config(width=7, height=1)

  # Configurando a posição do botão altera
  altera_button.place(relx=0.8, rely=0.7, anchor='center') 
  
  # Salvando a planilha
  wb.save(excel_file)
#Função responsável por alterar Status de uma OS
def Alterar_Status():
  # Criação e configuração da janela
  altera = tk.Toplevel()
  altera.title("Alterar Status")
  altera.geometry("600x95")
  altera.configure(background='#a8b5ab')

  # Caminho do arquivo Excel
  excel_file = 'MANUTENÇÕES INTERNAS WELLINGTON.xlsx'

  # Carrega o arquivo Excel
  wb = load_workbook(excel_file)
  planilha_radio = wb.active

  #Declaração de variáveis
  OS_Alterar_var = tk.StringVar()
  Status_alterar_var = tk.StringVar()

  # Criando as label
  OS_Alterar_label = tk.Label(altera, text = "OS procurada:", bg='#8a948c', font = Font_3)
  Status_alterar_label = tk.Label(altera, text = "Alterar Status:", bg='#8a948c', font = Font_3)

  # Posicionando os labels
  OS_Alterar_label.grid(row=0, column=0, sticky='E', pady=3)
  Status_alterar_label.grid(row=1, column=0, sticky='E', pady=3)

  # Criando o campo de entrada
  OS_Alterar_entry = tk.Entry(altera, textvariable = OS_Alterar_var, font = Font_2)

  # Posicionando o campo de entrada
  OS_Alterar_entry.grid(row=0, column=1, sticky='W')

  # Criando os frames para os botões de checagem
  frame_NovoStatus = tk.Frame(altera, bg='#a8b5ab')
  frame_NovoStatus.grid(row=1, column=1, sticky='W', padx=5)

  # Criando os botões de checagem
  Status_liberado = Radiobutton(frame_NovoStatus, text='Liberado', variable=Status_alterar_var,
    value='LIBERADO', bg='#a8b5ab', font=Font_2, fg='#01700e')
  status_PT = Radiobutton(frame_NovoStatus, text='PT', variable=Status_alterar_var,
  value='PT', bg='#a8b5ab', font=Font_2, fg='#e80707')            
  status_orçado = Radiobutton(frame_NovoStatus, text='Orçado', variable=Status_alterar_var,
  value='ORÇADO', bg='#a8b5ab', font=Font_2, fg='#ff4d01')   
  status_bancada = Radiobutton(frame_NovoStatus, text='Bancada', variable=Status_alterar_var,
   value='BANCADA', bg='#a8b5ab', fg='black', font=Font_2)
  stauts_AgPeca = Radiobutton(frame_NovoStatus, text='Aguardando Peça', variable=Status_alterar_var,
  value='AP', bg='#a8b5ab', font=Font_2, fg='#ffe200')

  # Posicionando os botões de checagem
  Status_liberado.pack(side='left', padx=5)
  status_PT.pack(side='left', padx=5)
  status_orçado.pack(side='left', padx=5)
  status_bancada.pack(side='left', padx=5)
  stauts_AgPeca.pack(side='left', padx=5)

  def alterando_campos():
    os_procurada = OS_Alterar_var.get()
    linha_encontrada = None

    # Itera sobre as linhas da planilha para encontrar a OS
    for row in planilha_radio.iter_rows(min_row=2, max_row=planilha_radio.max_row, min_col=1, max_col=1):
      for cell in row:
        if str(cell.value) == os_procurada:
          linha_encontrada = cell.row
          break

    if linha_encontrada:
      novo_status = Status_alterar_var.get()

      c8 = planilha_radio.cell(row=linha_encontrada, column=8)

      # Apaga os valores das células
      c8.value = None

      # Atribuindo os novos valores às células
      c8.value = novo_status

      # Mapeia as cores para os diferentes valores de Status
      cores = {'LIBERADO': 'FF92D050', 'PT': 'FF0000', 'AP': 'FFFF00', 'ORÇADO': 'FFBF00'}

      # Obtém a cor correspondente ao Status
      cor = cores.get(novo_status, 'FFFFFF')

      # Aplica a cor à célula 
      c8.fill = styles.PatternFill(start_color=cor, end_color=cor, fill_type='solid')

      # Salvando a planilha
      wb.save(excel_file)

    else:
      # Mostrando uma mensagem de erro se a OS não for encontrada
      messagebox.showinfo("Busca não encontrada", "O.S não encontrada")
      altera.destroy()

    # Zerando as strings
    OS_Alterar_var.set("")
    Status_alterar_var.set("")

  # Configurando um botão para que os novos dados sejam inseridos na planilha do excel
  altera_button = tk.Button(altera ,text = 'Alterar', command = alterando_campos)
  altera_button['background'] = '#667369'
  altera_button['activebackground'] = 'gray40'
  altera_button['fg'] = 'white'
  altera_button['font'] = Font_1
  altera_button.config(width=7, height=1)

  # Configurando a posição do botão altera
  altera_button.place(relx=0.85, rely=0.78, anchor='center') 

  # Salvando a planilha
  wb.save(excel_file)
#Função responsável por alterar Defeito de uma OS
def Alterar_Servico():
  # Criação e configuração da janela
  altera = tk.Toplevel()
  altera.title("Alterar Servico")
  altera.geometry("400x65")
  altera.configure(background='#a8b5ab')

  # Caminho do arquivo Excel
  excel_file = 'MANUTENÇÕES INTERNAS WELLINGTON.xlsx'

  # Carrega o arquivo Excel
  wb = load_workbook(excel_file)
  planilha_radio = wb.active

  #Declaração de variáveis
  OS_Alterar_var = tk.StringVar()
  Defeito_alterar_var = tk.StringVar()

  # Criando as label
  OS_Alterar_label = tk.Label(altera, text = "OS procurada:", bg='#8a948c', font = Font_3)
  Defeito_alterar_label = tk.Label(altera, text = "Adicionar Defeito:", bg='#8a948c', font = Font_3)

  # Posicionando os labels
  OS_Alterar_label.grid(row=0, column=0, sticky='E', pady=3)
  Defeito_alterar_label.grid(row=3, column=0, sticky='E', pady=3)

  # Criando o campo de entrada
  OS_Alterar_entry = tk.Entry(altera, textvariable = OS_Alterar_var, font = Font_2)
  Defeito_alterar_entry = tk.Entry(altera, textvariable = Defeito_alterar_var, font = Font_2)

  # Posicionando o campo de entrada
  OS_Alterar_entry.grid(row=0, column=1, sticky='W')
  Defeito_alterar_entry.grid(row=3, column=1, sticky='W')

  def alterando_campos():
    os_procurada = OS_Alterar_var.get()
    linha_encontrada = None

    # Itera sobre as linhas da planilha para encontrar a OS
    for row in planilha_radio.iter_rows(min_row=2, max_row=planilha_radio.max_row, min_col=1, max_col=1):
      for cell in row:
        if str(cell.value) == os_procurada:
          linha_encontrada = cell.row
          break

    if linha_encontrada:
      novo_defeito = Defeito_alterar_var.get()

      c12 = planilha_radio.cell(row=linha_encontrada, column=12)

      # Concatenando as strings
      valorAtual = c12.value
      Defeito_Concatenado = valorAtual + " "  + novo_defeito 
      
      # Atribuindo os novos valores às células
      c12.value = Defeito_Concatenado

      # Salvando a planilha
      wb.save(excel_file)

    else:
      # Mostrando uma mensagem de erro se a OS não for encontrada
      messagebox.showinfo("Busca não encontrada", "O.S não encontrada")
      altera.destroy()

    # Zerando as strings
    OS_Alterar_var.set("")
    Defeito_alterar_var.set("")

  # Configurando um botão para que os novos dados sejam inseridos na planilha do excel
  altera_button = tk.Button(altera ,text = 'Adicionar', command = alterando_campos)
  altera_button['background'] = '#667369'
  altera_button['activebackground'] = 'gray40'
  altera_button['fg'] = 'white'
  altera_button['font'] = Font_1
  altera_button.config(width=7, height=1)

  # Configurando a posição do botão altera
  altera_button.place(relx=0.8, rely=0.7, anchor='center') 

  # Salvando a planilha
  wb.save(excel_file)
# Função que da funcionalidade do botão submit
def Submit():

# Caminho do arquivo excel
  excel = 'MANUTENÇÕES INTERNAS WELLINGTON.xlsX'

  # Abrindo o arquivo excel
  wb = load_workbook(excel)

  # Atribuindo o arquivo excel a uma variável
  planilha_radio = wb.active

  # Atribuindo as entradas do usuario a outras variaveis para que possam ser armazenas no excel
  retorno = retorno_var.get()
  OrdemServicoAnt = OrdemServicoAnt_var.get()
  OrdemServico = OrdemServico_var.get()
  Garantia = Garantia_var.get()
  NumeroSerie = NumeroSerie_var.get()
  patrimonio =  Patrimonio_var.get()
  proprietario = Proprietario_var.get()
  equipamento = Equipamento_var.get()
  status = Status_var.get()
  idenizacao = Idenizacao_var.get()
  data = date.today()
  data_formatada = data.strftime('%d/%m/%Y')
  servico_executado = Servico_var.get()

  # Calcular a próxima linha disponível
  NovaLinha_inserida = 2

  # Inserir uma nova linha após a primeira
  planilha_radio.insert_rows(NovaLinha_inserida)
  
  # Dados a serem inseridos
  dados = [
      OrdemServico,
      patrimonio,
      OrdemServicoAnt,
      Garantia,
      NumeroSerie,
      proprietario,
      equipamento,
      status,
      idenizacao,
      data_formatada,
      retorno,
      servico_executado
  ]
  # Escrevendo os dados no excel
  for col, valor in enumerate(dados, start=1):
    planilha_radio.cell(row=NovaLinha_inserida, column=col, value=valor)

  # Colorir a célula correspondente ao Status
  cor_status = {'LIBERADO': 'FF92D050', 'PT': 'FF0000', 'AP': 'FFFF00', 'ORÇADO': 'FFBF00'}

  # Obtém a cor correspondente ao Status
  cor = cor_status.get(status, 'FFFFFF')

  # Aplica a cor à célula 
  planilha_radio.cell(row=NovaLinha_inserida, column=8).fill = styles.PatternFill(start_color=cor, end_color=cor, fill_type='solid')

  # Salvando a planilha
  wb.save(excel)

  # Zerando as strings
  retorno_var.set("")
  OrdemServicoAnt_var.set("")
  OrdemServico_var.set("")
  Garantia_var.set("")
  NumeroSerie_var.set("")
  Patrimonio_var.set("")
  Proprietario_var.set("")
  Equipamento_var.set("")
  Status_var.set("")
  Idenizacao_var.set("")
  data = None
  Servico_var.set("") 
# *  *  *  *  *  *  * #
#    FUNÇÃO MAIN      #
# *  *  *  *  *  *  * #    

# Criando a janela princial
radio = tk.Tk()
radio.geometry("625x350")
radio.title("Planilha Manutenções")
radio.configure(background='#a8b5ab')

# Criando as variaveis de entrada
retorno_var = tk.StringVar()
OrdemServico_var = tk.StringVar()
OrdemServicoAnt_var = tk.StringVar()
Garantia_var = tk.StringVar()
NumeroSerie_var = tk.StringVar()
Patrimonio_var = tk.StringVar()
Proprietario_var = tk.StringVar()
Equipamento_var = tk.StringVar()
Status_var = tk.StringVar()
Idenizacao_var = tk.StringVar()
Servico_var = tk.StringVar()

# Declarando as fontes
Font_1 = ("Helvetica", 10, "bold") 
Font_2 = ('Helvetica',10,)
Font_3 = ('Helvetica',11,'bold')

# Criando uma barra de menu
barra_menu = tk.Menu(radio, font=Font_1, background='#a8b5ab', fg='white')
radio.config(menu = barra_menu)
menu_versao = tk.Menu(barra_menu, tearoff=False)

# Criando o sobremenu 'Sobre'
barra_menu.add_cascade(label = "Versão", menu = menu_versao)
menu_versao.add_command(label = "Sobre", command = Sobre)
menu_versao.add_separator()
menu_versao.add_command(label = "Versão", command = Versao)

# Criando o sobremenu 'Localizar'
menu_localizar = tk.Menu(barra_menu, tearoff=False)
barra_menu.add_cascade(label = "Pesquisar", menu = menu_localizar)
menu_localizar.add_command(label = "Pesquisar", command = Localizar, accelerator="Ctrl+f")

# Criando o sobremenu 'Alterar'
menu_alterar = tk.Menu(barra_menu, tearoff=False)
barra_menu.add_cascade(label='Alterar', menu=menu_alterar)
menu_alterar.add_command(label='Status', command=Alterar_Status, accelerator="Ctrl+s")
menu_alterar.add_separator()
menu_alterar.add_command(label = "Data", command = Alterar_Data,  accelerator="Ctrl+d")
menu_alterar.add_separator()
menu_alterar.add_command(label = "Defeito", command = Alterar_Servico, accelerator="Ctrl+a")

# Criando o sobmenu 'Sair'
menu_sair = tk.Menu(barra_menu, tearoff=False)
barra_menu.add_cascade(label = "SAIR", menu = menu_sair)
menu_sair.add_command(label = "Sair", command = exit)

# Declarando as binds para atalhos do teclado
radio.bind_all("<Control-f>", lambda event: Localizar())
radio.bind_all("<Control-s>", lambda event: Alterar_Status())
radio.bind_all("<Control-d>", lambda event: Alterar_Data())
radio.bind_all("<Control-a>", lambda event: Alterar_Servico())

# Criando as opções de entry dentro da janela principal
OS_entry = tk.Entry(radio, textvariable = OrdemServico_var, font = Font_2)
OS_Anterior_entry = tk.Entry(radio, textvariable = OrdemServicoAnt_var, font =Font_2)
NumeroSerie_entry = tk.Entry(radio, textvariable = NumeroSerie_var, font = Font_2)
Patrimonio_entry = tk.Entry(radio, textvariable = Patrimonio_var, font = Font_2)
Equipamento_entry = tk.Entry(radio, textvariable = Equipamento_var, font = Font_2)
Servico_entry = tk.Entry(radio, textvariable = Servico_var, font = Font_2)

# Configurando a posição das entry dentro da janela principal
OS_entry.grid(row=0, column=1, sticky='W')
OS_Anterior_entry.grid(row=2, column=1, sticky='W')
NumeroSerie_entry.grid(row=4, column=1, sticky='W')
Patrimonio_entry.grid(row=5, column=1, sticky='W')
Equipamento_entry.grid(row=7, column=1, sticky='W')
Servico_entry.grid(row=10,column=1, sticky='W')

# Criando as label das entry dentro da janela principal
Retorno_label = tk.Label(radio, text = "Retorno:", bg='#8a948c', font = Font_3)
Garantia_label = tk.Label(radio, text = "Garantia:", bg='#8a948c', font = Font_3)
Proprietario_label = tk.Label(radio, text = "Proprietário:", bg='#8a948c', font = Font_3)
Status_label = tk.Label(radio, text = "Status:", bg='#8a948c', font = Font_3)
Idenizacao_label = tk.Label(radio, text = "Idenização:", bg='#8a948c', font = Font_3)
OS_label = tk.Label(radio, text = 'O.S', bg='#8a948c', font = Font_3)
OS_Anterior_label = tk.Label(radio, text = 'O.S Anterior:', bg='#8a948c', font = Font_3)
NumeroSerie_label = tk.Label(radio, text = 'Numero Serie:', bg='#8a948c', font = Font_3)
Patrimonio_label = tk.Label(radio, text = 'Patrimonio:', bg='#8a948c', font = Font_3)
Equipamento_label = tk.Label(radio, text = 'Equipamento:', bg='#8a948c', font = Font_3)
Servico_label = tk.Label(radio, text = 'Defeito:', bg='#8a948c', font = Font_3)

# Configurando a posição das label dentro da janela principal
Retorno_label.grid(row=1, column=0,pady=3)
OS_label.grid(row=0, column=0,pady=3)
OS_Anterior_label.grid(row=2, column=0, pady=3)
Garantia_label.grid(row=3, column=0, pady=3)
NumeroSerie_label.grid(row=4, column=0, pady=3)
Patrimonio_label.grid(row=5, column=0, pady=3)
Proprietario_label.grid(row=6, column=0, pady=3)
Equipamento_label.grid(row=7, column=0, pady=3)
Status_label.grid(row=8, column=0, pady=3)
Idenizacao_label.grid(row=9, column=0,pady=3)
Servico_label.grid(row=10,column=0, pady=3)

# Configurando um botão para que os dados sejam inseridos na planilha do excel
submit_button = tk.Button(radio ,text = 'Submit', command = Submit)
submit_button['background'] = '#667369'
submit_button['activebackground'] = 'gray40'
submit_button['fg'] = 'white'
submit_button['font'] = Font_1
submit_button.config(width=13, height=3)

# Configurando a posição do botão submit
submit_button.place(relx=0.8, rely=0.88, anchor='center') 

# Criando os frames para cada botão de chegagem
frame_retorno = tk.Frame(radio, bg='#a8b5ab')
frame_retorno.grid(row=1, column=1, sticky='W', padx=5)

frame_garantia = tk.Frame(radio, bg='#a8b5ab')
frame_garantia.grid(row=3, column=1, sticky='W', padx=5)

frame_proprietario = tk.Frame(radio, bg='#a8b5ab')
frame_proprietario.grid(row=6, column=1, sticky='W', padx=5)

frame_status = tk.Frame(radio, bg='#a8b5ab')
frame_status.grid(row=8, column=1, sticky='W', padx=5)

frame_idenizacao = tk.Frame(radio, bg='#a8b5ab')
frame_idenizacao.grid(row=9, column=1, sticky='W', padx=5)

# Criando os botões de chegagem 
retorno_var = StringVar()
Retorno_S = Radiobutton(frame_retorno, text="Sim", variable=retorno_var, 
                          value="SIM", bg='#a8b5ab', fg='#01700e', font=Font_2)
Retorno_N = Radiobutton(frame_retorno, text="Não", variable=retorno_var, 
                          value="NÃO", bg='#a8b5ab', fg='#e80707', font=Font_2)

GarantiaRompida_FAB = Radiobutton(frame_garantia, text="Fabrica", variable=Garantia_var,
                                  value="FABRICA", bg='#a8b5ab', fg='#8103ff', font=Font_2)
GarantiaRompida_RUIM = Radiobutton(frame_garantia, text="Ruim", variable=Garantia_var,
                                  value="RUIM", bg='#a8b5ab', fg='#e80707', font=Font_2)

Proprietario_loc = Radiobutton(frame_proprietario, text='Locação', variable=Proprietario_var,
                                 value='LOCAÇÃO', bg='#a8b5ab', fg='#8f310e', font=Font_2)
Proprietario_cliente = Radiobutton(frame_proprietario, text='Cliente', variable=Proprietario_var,
                                     value='CLIENTE', bg='#a8b5ab', fg='#071eed', font=Font_2)

Status_liberado = Radiobutton(frame_status, text='Liberado', variable=Status_var,
                                value='LIBERADO', bg='#a8b5ab', font=Font_2, fg='#01700e')
status_PT = Radiobutton(frame_status, text='PT', variable=Status_var,
                          value='PT', bg='#a8b5ab', font=Font_2, fg='#e80707')            
status_orçado = Radiobutton(frame_status, text='Orçado', variable=Status_var,
                              value='ORÇADO', bg='#a8b5ab', font=Font_2, fg='#ff4d01')   
status_bancada = Radiobutton(frame_status, text='Bancada', variable=Status_var,
                               value='BANCADA', bg='#a8b5ab', fg='black', font=Font_2)
stauts_AgPeca = Radiobutton(frame_status, text='Aguardando Peça', variable=Status_var,
                              value='AP', bg='#a8b5ab', font=Font_2, fg='#ffe200')

idenizacao_S = Radiobutton(frame_idenizacao, text='Sim', variable=Idenizacao_var,
                             value='SIM', bg='#a8b5ab', font=Font_2, fg='#01700e')
idenizacao_N = Radiobutton(frame_idenizacao, text='Não', variable=Idenizacao_var,
                             value='NÃO', bg='#a8b5ab', font=Font_2, fg='#e80707')

# Configurando a posição dos botões de checagem dentro da janela principal
Retorno_S.pack(side='left', padx=5)
Retorno_N.pack(side='left', padx=5)

GarantiaRompida_FAB.pack(side='left', padx=5)
GarantiaRompida_RUIM.pack(side='left', padx=5)

Proprietario_loc.pack(side='left', padx=5)
Proprietario_cliente.pack(side='left', padx=5)

Status_liberado.pack(side='left', padx=5)
status_PT.pack(side='left', padx=5)
status_orçado.pack(side='left', padx=5)
status_bancada.pack(side='left', padx=5)
stauts_AgPeca.pack(side='left', padx=5)

idenizacao_S.pack(side='left', padx=5)
idenizacao_N.pack(side='left', padx=5)

# Loop para a jenla princial
radio.mainloop()
